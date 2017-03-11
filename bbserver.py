#coding:utf-8
import sys
import os
reload(sys)
sys.setdefaultencoding("utf-8")
from flask import Flask
from flask import request
from flask import send_from_directory
from flask import url_for
from flask import render_template
from flask import redirect
import urllib
import urllib2
from urllib2 import urlopen
import ssl
import json
from openpyxl import load_workbook
import data_fetch

context = ssl.SSLContext(ssl.PROTOCOL_TLSv1_2)
context.load_cert_chain('server.crt', 'server.key')
app = Flask(__name__, static_url_path='')

@app.route('/')
def index():
    #return 'test'
    base_url = 'https://login.microsoftonline.com/common/oauth2/v2.0/authorize?'
    redirect_uri = 'https://bb.xjq314.com/code'
    client_secret = 'xdv5xM0og534xL6fnRRXhDe'
    client_id='e194ad94-2aa1-4b8f-b299-0e0dd2df136d'
    scopes='files.read files.read.all'
    response_type = 'code'
    f = {
        'client_id': client_id,
        'scope': scopes,
        'response_type': response_type,
        'redirect_uri': redirect_uri
    }
    # Client requests an authorization token
    url = base_url + urllib.urlencode(f)
    return redirect(url)

@app.route('/code')
def code():
    code = request.args.get('code', '')
    if code == '':
        return 'get code fail'
    else:
        base_url = 'https://login.microsoftonline.com/common/oauth2/v2.0/token'
        client_id = 'e194ad94-2aa1-4b8f-b299-0e0dd2df136d'
        redirect_uri = 'https://bb.xjq314.com/code'
        client_secret = 'n5b0Hf1TNomMgqkJpzbbb10'
        grant_type = 'authorization_code'
        f = {
            'client_id': client_id,
            'redirect_uri': redirect_uri,
            'client_secret' : client_secret,
            'code' : code,
            'grant_type' : grant_type
        }
        data = urllib.urlencode(f)
        headers = {
                  'Content-Type': 'application/x-www-form-urlencoded'
        }
        # Client redeems the auth code for an access token
        access_token = _get_access_token(code)
        # Cilent download file through pre-auth url
        url = _get_file_url(access_token)
        _download_file(url)
        wb2 = load_workbook('training.xlsx')
        return redirect('/current')

@app.route('/data')
def data():
    wb = load_workbook('training.xlsx')
    body_weight_list = data_fetch.get_body_weight(wb, 'BMR')
    up_actions_dict = data_fetch.get_up_actions_weight(wb)
    down_actions_dict = data_fetch.get_down_actions_weight(wb)
    data = {
        'weight': body_weight_list,
        'up': up_actions_dict,
        'down': down_actions_dict 
    }
    return json.dumps(data)

@app.route('/current')
def test():
    return render_template('basic.html')

@app.route('/css/<filename>')
def css(filename):
    return send_from_directory('static/css', filename)

@app.route('/js/<filename>')
def js(filename):
    return send_from_directory('static/js', filename)

@app.route('/fonts/<filename>')
def fonts(filename):
    return send_from_directory('static/fonts', filename)

def _download_file(url):
    res = urllib2.urlopen(url)
    with open('training.xlsx', 'w') as local_file:
        local_file.write(res.read())

def _get_file_url(access_token):
    # Cilent download file through pre-auth url
    headers = {
                  'Authorization': 'bearer ' + access_token
        }
    url = 'https://graph.microsoft.com/v1.0/drive/items/8CC5F2122D1F44F8!521'
    req = urllib2.Request(url, headers=headers)
    res = urllib2.urlopen(req)
    data = json.load(res)
    url = data['@microsoft.graph.downloadUrl']
    return url

def _get_access_token(code):
    # Client redeems the auth code for an access token
    base_url = 'https://login.microsoftonline.com/common/oauth2/v2.0/token'
    client_id = 'e194ad94-2aa1-4b8f-b299-0e0dd2df136d'
    redirect_uri = 'https://bb.xjq314.com/code'
    client_secret = 'n5b0Hf1TNomMgqkJpzbbb10'
    grant_type = 'authorization_code'
    f = {
        'client_id': client_id,
        'redirect_uri': redirect_uri,
        'client_secret' : client_secret,
        'code' : code,
        'grant_type' : grant_type
    }
    data = urllib.urlencode(f)
    headers = {
              'Content-Type': 'application/x-www-form-urlencoded'
    }
    url = base_url
    req = urllib2.Request(url, data=data, headers=headers)
    req.get_method = lambda: 'POST'
    opener = urllib2.build_opener()
    res = opener.open(req)
    data = json.load(res)
    token_type = data['token_type']
    expires_in = data['expires_in']
    scope = data['scope']
    access_token = data['access_token']
    return access_token


if __name__ == '__main__':
    app.debug = True
    app.run(host='10.0.0.4', port=443, threaded=True, ssl_context=context)
